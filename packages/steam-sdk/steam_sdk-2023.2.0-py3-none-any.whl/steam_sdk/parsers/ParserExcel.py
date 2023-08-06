import os
import datetime

import numpy as np
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from steam_sdk.utils.make_folder_if_not_existing import make_folder_if_not_existing


def read_row(workSheet, Nrow, St=False):
    '''
        *** Function that reads a row of a worksheet of an Excel file ***

        :param workSheet: Name of the worksheet
        :type workSheet: str
        :param Nrow: Number of the row
        :type Nrow: int
        :param St: Flag indicating whether the entry to read is a string
        :type St: bool

        :return: np.array([])
    '''
    rowValues = np.array([])
    row = workSheet[Nrow]
    for cell in row:
        if not St:
            if isinstance(cell.value, str): continue
        rowValues = np.append(rowValues, cell.value)
    rowValues = rowValues[rowValues != None]
    return rowValues


def write2Excel(name_file: str,
                name_sheets: list = ['Sheet1'],
                listOf_variables_values_descriptions: list = [[['test_name_1', 'test_value_1', 'test_description_1'], [None, None, 'test_title'], ['test_name_2', 'test_value_2', 'test_description_2']]],
                verbose: bool = False):
    '''
        *** Function that writes an Excel file ***

        Function to write an Excel file composed of multiple sheets

        :param name_file: String defining the name of the .xlsx input file to be written
        :type name_file: string
        # TODO Add entries
        :param verbose: flag that determines whether the output are printed
        :type verbose: bool

        :return: None
    '''
    # TODO: description more general

    # If the output folder does not exist, make it now
    make_folder_if_not_existing(os.path.dirname(name_file))

    # In case a string is passed, make it a 1-element list
    if isinstance(name_sheets, str):
        name_sheets = [name_sheets]

    # Start an empty workbook
    workbook = openpyxl.Workbook()
    workbook.properties.creator = 'STEAM-Team'

    # Sequentially write the sheets, one by one
    for sheet in range(len(listOf_variables_values_descriptions)):
        variables_values_descriptions = listOf_variables_values_descriptions[sheet]
        if verbose:
            print('### Write sheet #{} named "{}" ###'.format(sheet+1, name_sheets[sheet]))
            print('variables_values_descriptions{} = '.format(variables_values_descriptions))
            print('len(variables_values_descriptions) = {}'.format(len(variables_values_descriptions)))
        writeWorkbookSheet(workbook, name_sheets[sheet], variables_values_descriptions, verbose)

    # Remove the sheet initially added by default
    std = workbook['Sheet']
    workbook.remove(std)

    # Set the active sheet as the first one in the sheet list
    select_active_sheet(workbook, name_sheets[0])

    # Save the workbook
    workbook.save(name_file)

    # Display time stamp
    currentDT = datetime.datetime.now()
    if verbose:
        print(' ')
        print('Time stamp: ' + str(currentDT))
        print('New file ' + name_file + ' generated.')
    return workbook


def select_active_sheet(workbook, name_sheet):
    '''
    ** Select active sheet in a workbook **

    :param workbook:
    :param name_sheet:
    :return: sheet
    #TODO add descriptions
    '''

    for s in range(len(workbook.sheetnames)):
        if workbook.sheetnames[s] == name_sheet:
            break
    workbook.active = s
    return s

def writeWorkbookSheet(workbook, name_sheet, variables_values_descriptions, verbose: bool=False):
    """
        **Write one sheet of an Excel file**

        Function writes one sheet of an Excel file

        #TODO Add description
        :return:
    """

    # Define cell style
    def style_title_cell(data):
        for c in data:
            c = Cell(active_sheet, column="A", row=1, value=c)
            c.font = Font(size=14, bold=True)
            yield c

    # Loop through sheets to activate the correct sheet to write to
    workbook.create_sheet(index=len(workbook.worksheets) + 1, title=name_sheet)
    for s in range(len(workbook.sheetnames)):
        if workbook.sheetnames[s] == name_sheet: break
    workbook.active = s
    active_sheet = workbook.active

    # Write entries in the sheet, row by row
    for r, row in enumerate(variables_values_descriptions):
        name, value, description = row[0], row[1], row[2]
        if verbose:
            print('row: {}'.format(row))
        # Check if size of value list/array is > 16382. If so, convert to vertical np.ndarray
        # Note: Max number of columns in an .xlsx file is 16384, and two columns are needed for variable name and description.
        if isinstance(value, np.ndarray) and value.shape[0] > 16382:
            value = np.array(value).reshape(-1, 1)
            if verbose:
                print('row #{} has more than 16382 elements and will be written as a column array.'.format(r+1))
        if isinstance(value, list) and len(value) > 16382:
            value = np.array(value).reshape(-1, 1)
            if verbose:
                print('row #{} has more than 16382 elements and will be written as a column array.'.format(r+1))

        # Special case when row is an empty row
        if name == None and value == None and description == None:
            if verbose:
                print('row #{} will be treated as an empty row.'.format(r+1))
            active_sheet.append([None, None, None])  # Write empty row
            continue  # stop treating this row: go to the next row

        # Special case when row is a title
        if name == None and value == None:
            if verbose:
                print('row #{} will be treated as a title.'.format(r+1))
            active_sheet.append(style_title_cell([description]))  # Write title in the first column, with special style (bold text, larger font)
            continue  # stop treating this row: go to the next row

        # Write the variable. Check which datatype the variable value is and append it to the sheet
        # If datatype is a list of lists, a for loop executes all rows/columns before continuing to next variable
        if isinstance(value, list) and any(isinstance(el, list) for el in value):
            for i in range(len(value)):
                values = value[i]
                if i == 0:
                    active_sheet.append([description, name] + values)  # first row: contains also variable description and name
                else:
                    active_sheet.append([None, None] + values)         # other rows: contain only value
            continue  # stop treating this row: go to the next row
        # If datatype is a matrix, a for loop executes all rows/columns before continuing to next variable
        if isinstance(value, np.ndarray) and value.ndim > 1:
            for i in range(value.shape[0]):
                values = value[i, :].tolist()
                if i == 0:
                    active_sheet.append([description, name] + values)  # first row: contains also variable description and name
                else:
                    active_sheet.append([None, None] + values)         # other rows: contain only value
            continue  # stop treating this row: go to the next row
        # If datatype is an np.array, make it a list
        if isinstance(value, np.ndarray):
            value = np.array(value).tolist()
        # If datatype is not a list, make it a list
        if not isinstance(value, list):
            value = [value]
        # Write the variable
        active_sheet.append([description, name] + value)


        # Setting the width of each cells in the workbook [only for good view]
        # TODO Assign width to the correct column: form A, B, etc to column_numbers[0]...
        width = [80.7109375, 40.7109375, 20.7109375]
        active_sheet.column_dimensions['A'].width = width[0]
        active_sheet.column_dimensions['B'].width = width[1]

        # TODO Assign width to all other columns
        if active_sheet.max_column + 1 > 18278:
            smc = 18278
        else:
            smc = active_sheet.max_column + 1
        for i in range(3, smc):
            cl = get_column_letter(i)
            active_sheet.column_dimensions[cl].width = width[2]


def compare_two_parameters(var_1, var_2, Diff, attribute, max_relative_error, show_indices):
    '''
        Helper-function for ParserLEDET() and ParserProteCCT() to compare to parameters
        max_relative_error is a relative value
    '''
    Block = 1

    # Check whether there are parameters in one file and not in the other
    if var_1 is not None and var_2 is None:
        Diff = 1
        print('Parameter {} in file A is {} while in file B is {}'.format(attribute, var_1, var_2))
        return Diff

    if var_2 is not None and var_1 is None:
        Diff = 1
        print('Parameter {} in file B is {} while in file A is {}'.format(attribute, var_2, var_1))
        return Diff

    # Check parameter differences
    if isinstance(var_1, float) or isinstance(var_1, int):
        if abs(var_1 - var_2) > max_relative_error * abs(var_1):
            if Block: print(
                "Found difference in scalar Parameter {}, A: {}, B: {}".format(attribute, var_1, var_2))
            Block = 0
            Diff = True

    elif var_1 is None or var_2 is None:
        if var_1 is not None or var_2 is not None:
            Diff = 1
            print('Parameter {} in file A is {} while in file B is {}'.format(attribute, var_1, var_2))

    elif len(var_1) != len(var_2):
        Diff = 1
        if Block:
            Block = 0
            print('Parameter {} of A, {} has not the same length as Parameter of B, {}'.format(attribute,
                                                                                               len(var_1),
                                                                                               len(var_2)))
    else:
        Pos = []
        for k in range(len(var_1)):
            try:
                if var_1[k] != var_2[k]:
                    if isinstance(var_1[k], str):
                        if var_1[k] != var_2[k]:
                            Diff = 1
                            if Block:
                                print("Found difference in vector Parameter {}".format(attribute))
                                Block = 0
                            Pos.append(k)

                    elif abs(var_1[k] - var_2[k]) > max_relative_error * abs(var_1[k]):
                        Diff = 1
                        if Block:
                            print("Found difference in vector Parameter {}".format(attribute))
                            Block = 0
                        Pos.append(k)
            except:
                for j in range(var_1.shape[1]):
                    if var_1[k, j] != var_2[k, j]:
                        if isinstance(var_1[k], str):
                            Diff = 1
                            if Block:
                                print("Found difference in matrix Parameter {}".format(attribute))
                                Block = 0
                            Pos.append([k, j])
                        elif abs(var_1[k, j] - var_2[k, j]) > max_relative_error * abs(var_1[k, j]):
                            Diff = 1
                            if Block:
                                print("Found difference in matrix Parameter {}".format(attribute))
                                Block = 0
                            Pos.append([k, j])

        if len(Pos) > 0:
            if len(Pos) < 10:
                print("Different Positions: {}".format(Pos))
            else:
                print("Many values are different (>10)")
                if show_indices: print(Pos)

    return Diff