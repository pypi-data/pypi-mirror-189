import re
import time
import datetime
import requests
import numpy as np
import  pandas as pd
import openpyxl.utils.cell as c
from bs4 import BeautifulSoup
from selenium import webdriver
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from pandas import DataFrame,Series
from win32com.client import Dispatch
from openpyxl.formula.translate import Translator


def find_income(company_id, n, google_driver_path, com_revenue_monthly_path, date_col, rev_col):
    driver = webdriver.Chrome(executable_path=google_driver_path)
    driver.get("https://mops.twse.com.tw/mops/web/t05st10_ifrs")
    time.sleep(1)
    input_from_list = driver.find_element_by_id("co_id").send_keys(str(company_id))
    time.sleep(1)
    button = driver.find_element_by_css_selector('[value=" 查詢 "]')
    button.click()
    windows = driver.window_handles
    driver.switch_to.window(windows[-1])
    time.sleep(1.5)
    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")
    driver.close()
    table = soup.find_all("div", {"id": "table01"})
    for t in table:
        tr_tag = t.find_all('tr')
    if tr_tag == []:
        find_income(company_id, n, google_driver_path, com_revenue_monthly_path, date_col, rev_col)
    else:
        new_income = tr_tag[4].text[3:].split()
        new_date = tr_tag[2].text[:9].split()
        if new_date[0][:2] == '民國':
            new_date = new_date
        else:
            new_date = tr_tag[1].text[:9].split()
        wb = load_workbook(filename=com_revenue_monthly_path)
        ws = wb.active
        ws._current_row = n
        ws.append({rev_col: new_income[0]})
        ws._current_row = n
        ws.append({date_col: new_date[0]})
        wb.save(filename=com_revenue_monthly_path)


def find_row(company_name, sheetname):
    company_row = 'None'
    for col in sheetname.columns:
        for row in col:
            if row.value == company_name:
                company_row = str(row.row)
    return company_row


def find_col(date, sheetname):
    for row in sheetname.rows:
        for col in row:
            if col.value == date:
                date_col = c.get_column_letter(col.col_idx)
                return date_col


def find_address(com_name, date, sheetname):
    if find_col(date, sheetname) != 'None' and find_row(com_name, sheetname) != 'None':
        address = find_col(date, sheetname) + find_row(com_name, sheetname)
    else:
        address = "None"
    return address


def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()

def twweb_to_excel(num,google_driver_path,com_revenue_monthly_path,date_col, rev_col):
    wb = load_workbook(com_revenue_monthly_path)
    ws = wb.active
    v = []
    for row in ws.rows:
        v.append(row[0].value)
    for i in range(num,179):
        find_income(v[i][:4], i, google_driver_path,com_revenue_monthly_path,date_col, rev_col)
        print(i)